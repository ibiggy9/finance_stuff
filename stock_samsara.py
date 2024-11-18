import pandas as pd
import datetime

# Create the data list
data = [
    ["Nov 18, 2024", 49.85, 50.27, 49.30, 49.31, 49.31, 804623],
    ["Nov 1, 2024", 48.00, 53.98, 45.76, 49.51, 49.51, 28698600],
    ["Oct 1, 2024", 48.01, 51.53, 45.31, 47.79, 47.79, 49948300],
    ["Sep 1, 2024", 40.73, 50.07, 37.62, 48.12, 48.12, 101365900],
    ["Aug 1, 2024", 38.11, 42.49, 32.22, 41.07, 41.07, 62396300],
    ["Jul 1, 2024", 33.60, 39.21, 33.22, 38.28, 38.28, 58210500],
    ["Jun 1, 2024", 34.50, 34.97, 27.14, 33.70, 33.70, 119214600],
    ["May 1, 2024", 34.92, 42.28, 32.67, 33.93, 33.93, 67432200],
    ["Apr 1, 2024", 37.60, 37.89, 29.70, 34.93, 34.93, 65057200],
    ["Mar 1, 2024", 34.86, 40.54, 32.25, 37.79, 37.79, 111951100],
    ["Feb 1, 2024", 31.68, 36.78, 30.71, 34.55, 34.55, 68218300],
    ["Jan 1, 2024", 32.75, 35.18, 29.80, 31.40, 31.40, 57394000],
    ["Dec 1, 2023", 31.00, 36.91, 29.63, 33.38, 33.38, 97125000],
    ["Nov 1, 2023", 22.86, 28.71, 21.48, 27.54, 27.54, 53815100],
    ["Oct 1, 2023", 25.22, 26.24, 21.76, 23.07, 23.07, 45259900],
    ["Sep 1, 2023", 27.00, 32.41, 22.70, 25.21, 25.21, 83309500],
    ["Aug 1, 2023", 27.68, 28.10, 21.89, 27.36, 27.36, 52963000],
    ["Jul 1, 2023", 27.75, 29.79, 25.06, 27.94, 27.94, 51504300],
    ["Jun 1, 2023", 19.01, 30.91, 18.61, 27.71, 27.71, 121585900],
    ["May 1, 2023", 18.19, 20.91, 16.63, 19.25, 19.25, 52344900],
    ["Apr 1, 2023", 19.74, 22.59, 17.67, 18.05, 18.05, 52548800],
    ["Mar 1, 2023", 16.68, 21.49, 16.08, 19.72, 19.72, 115901200],
    ["Feb 1, 2023", 13.73, 16.82, 13.35, 16.66, 16.66, 40994100],
    ["Jan 1, 2023", 12.55, 14.12, 10.48, 13.64, 13.64, 18929900],
    ["Dec 1, 2022", 9.55, 14.45, 9.44, 12.43, 12.43, 35506700],
    ["Nov 1, 2022", 12.57, 12.68, 8.42, 9.53, 9.53, 32171400],
    ["Oct 1, 2022", 12.24, 13.86, 10.58, 12.31, 12.31, 20102600],
    ["Sep 1, 2022", 14.86, 15.00, 11.33, 12.07, 12.07, 37321600],
    ["Aug 1, 2022", 14.34, 17.49, 14.06, 14.87, 14.87, 24741600],
    ["Jul 1, 2022", 11.21, 15.32, 11.03, 14.46, 14.46, 27676300],
    ["Jun 1, 2022", 11.09, 13.05, 9.86, 11.17, 11.17, 46705200],
    ["May 1, 2022", 12.24, 13.12, 8.72, 11.25, 11.25, 32586300],
    ["Apr 1, 2022", 15.85, 17.65, 12.30, 12.34, 12.34, 23365800],
    ["Mar 1, 2022", 17.42, 19.45, 12.51, 16.02, 16.02, 38735800],
    ["Feb 1, 2022", 18.10, 25.42, 15.03, 17.48, 17.48, 29171000],
    ["Jan 1, 2022", 28.75, 29.10, 14.55, 18.10, 18.10, 35491300],
    ["Dec 1, 2021", 33.65, 35.82, 27.15, 29.08, 29.08, 29405700],
    ["Nov 1, 2021", 43.11, 45.85, 31.54, 33.53, 33.53, 32641400],
    ["Oct 1, 2021", 41.87, 44.71, 40.80, 43.20, 43.20, 17613900],
    ["Sep 1, 2021", 44.25, 45.09, 39.30, 41.82, 41.82, 22866800],
    ["Aug 1, 2021", 39.85, 44.57, 38.10, 44.07, 44.07, 24714300],
    ["Jul 1, 2021", 40.88, 42.25, 36.25, 39.69, 39.69, 23741500],
    ["Jun 1, 2021", 38.61, 42.89, 37.29, 40.94, 40.94, 35762400],
    ["May 1, 2021", 38.58, 39.95, 30.27, 38.44, 38.44, 41726800],
    ["Apr 1, 2021", 35.59, 39.10, 34.32, 38.63, 38.63, 31011600],
    ["Mar 1, 2021", 36.44, 38.96, 31.45, 35.21, 35.21, 41990800],
    ["Feb 1, 2021", 37.03, 41.97, 35.54, 36.13, 36.13, 31228400],
    ["Jan 1, 2021", 33.00, 38.96, 32.58, 37.04, 37.04, 31834800],
    ["Dec 1, 2020", 29.30, 33.77, 28.93, 32.90, 32.90, 27133300],
    ["Nov 1, 2020", 23.41, 29.44, 22.65, 29.29, 29.29, 31424800],
    ["Oct 1, 2020", 25.75, 26.07, 22.02, 23.11, 23.11, 23387500],
    ["Sep 1, 2020", 29.01, 29.32, 21.84, 25.67, 25.67, 35614500],
    ["Aug 1, 2020", 24.24, 29.27, 24.11, 28.79, 28.79, 28509900],
    ["Jul 1, 2020", 23.37, 26.92, 22.85, 24.38, 24.38, 31287700],
    ["Jun 1, 2020", 23.09, 25.46, 21.70, 23.37, 23.37, 39531400],
    ["May 1, 2020", 19.00, 23.66, 18.52, 23.09, 23.09, 35099400],
    ["Apr 1, 2020", 16.15, 20.42, 15.80, 19.00, 19.00, 32271600],
    ["Mar 1, 2020", 21.77, 22.35, 14.56, 16.19, 16.19, 49195700],
    ["Feb 1, 2020", 23.85, 26.09, 19.46, 21.73, 21.73, 27632600],
    ["Jan 1, 2020", 21.48, 24.69, 21.40, 23.87, 23.87, 23395700]
]

# Create a DataFrame
df = pd.DataFrame(data, columns=['Date', 'Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume'])

# Convert Date string to datetime
df['Date'] = pd.to_datetime(df['Date'])

# Create Excel writer object with xlsm format
writer = pd.ExcelWriter('stock_data.xlsm', engine='openpyxl')

# Write DataFrame to Excel
df.to_excel(writer, sheet_name='Stock Data', index=False)

# Get the workbook and the worksheet
workbook = writer.book
worksheet = writer.sheets['Stock Data']

# Add formatting
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule

# Format headers
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Format number columns
for col in ['B', 'C', 'D', 'E', 'F']:  # Price columns
    for cell in worksheet[col]:
        if cell.row > 1:  # Skip header
            cell.number_format = '#,##0.00'

# Format volume column
for cell in worksheet['G']:
    if cell.row > 1:  # Skip header
        cell.number_format = '#,##0'

# Add color scaling to Close price column
worksheet.conditional_formatting.add('E2:E'+str(worksheet.max_row),
    ColorScaleRule(start_type='min', start_color='FF8080',
                  mid_type='percentile', mid_value=50, mid_color='FFEB84',
                  end_type='max', end_color='63BE7B'))

# Add borders
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

for row in worksheet.iter_rows():
    for cell in row:
        cell.border = thin_border

# Adjust column widths
for column in worksheet.columns:
    max_length = 0
    column = list(column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

# Save the workbook
workbook.save('stock_data.xlsx')