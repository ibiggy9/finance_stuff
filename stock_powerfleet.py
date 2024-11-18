import pandas as pd
import datetime

# Create the data list
data = [
    ["Nov 18, 2024", 5.95, 6.36, 5.88, 6.15, 6.15, 892196],
    ["Nov 1, 2024", 5.03, 6.68, 4.75, 5.95, 5.95, 16927300],
    ["Oct 1, 2024", 4.99, 5.81, 4.60, 5.03, 5.03, 22029500],
    ["Sep 1, 2024", 4.98, 5.53, 4.34, 5.00, 5.00, 20194600],
    ["Aug 1, 2024", 4.44, 5.22, 3.99, 4.99, 4.99, 19699200],
    ["Jul 1, 2024", 4.61, 5.25, 4.01, 4.53, 4.53, 15085000],
    ["Jun 1, 2024", 5.37, 5.38, 4.11, 4.57, 4.57, 28163300],
    ["May 1, 2024", 4.70, 5.58, 4.70, 5.31, 5.31, 18202500],
    ["Apr 1, 2024", 5.58, 5.67, 3.85, 4.79, 4.79, 19017500],
    ["Mar 1, 2024", 3.21, 5.49, 2.88, 5.34, 5.34, 23227500],
    ["Feb 1, 2024", 3.22, 3.27, 2.85, 3.17, 3.17, 2479600],
    ["Jan 1, 2024", 3.43, 3.43, 2.86, 3.20, 3.20, 3112000],
    ["Dec 1, 2023", 2.31, 3.43, 2.22, 3.42, 3.42, 3086200],
    ["Nov 1, 2023", 1.81, 2.45, 1.64, 2.32, 2.32, 4086300],
    ["Oct 1, 2023", 2.06, 3.25, 1.76, 1.85, 1.85, 14121000],
    ["Sep 1, 2023", 2.49, 2.49, 1.95, 2.07, 2.07, 1959000],
    ["Aug 1, 2023", 2.71, 2.73, 1.85, 2.46, 2.46, 1960500],
    ["Jul 1, 2023", 2.93, 3.06, 2.57, 2.71, 2.71, 1015600],
    ["Jun 1, 2023", 3.09, 3.45, 2.79, 3.00, 3.00, 1814900],
    ["May 1, 2023", 2.81, 3.20, 2.67, 3.13, 3.13, 3299500],
    ["Apr 1, 2023", 3.46, 3.49, 2.66, 2.87, 2.87, 773100],
    ["Mar 1, 2023", 2.76, 3.48, 2.56, 3.43, 3.43, 1841400],
    ["Feb 1, 2023", 2.91, 3.00, 2.52, 2.78, 2.78, 1690300],
    ["Jan 1, 2023", 2.75, 2.97, 2.50, 2.86, 2.86, 1031700],
    ["Dec 1, 2022", 2.78, 2.91, 2.25, 2.69, 2.69, 1218100],
    ["Nov 1, 2022", 2.59, 3.20, 2.40, 2.82, 2.82, 1098900],
    ["Oct 1, 2022", 3.10, 3.29, 2.65, 2.65, 2.65, 1892300],
    ["Sep 1, 2022", 3.18, 3.33, 2.75, 3.08, 3.08, 538100],
    ["Aug 1, 2022", 2.85, 3.97, 2.55, 3.19, 3.19, 1108400],
    ["Jul 1, 2022", 2.17, 3.05, 2.06, 2.75, 2.75, 697700],
    ["Jun 1, 2022", 2.30, 2.61, 2.13, 2.17, 2.17, 1746100],
    ["May 1, 2022", 2.54, 2.94, 2.16, 2.37, 2.37, 1620100],
    ["Apr 1, 2022", 2.98, 3.12, 2.55, 2.64, 2.64, 1304800],
    ["Mar 1, 2022", 3.53, 3.70, 2.68, 2.97, 2.97, 2495200],
    ["Feb 1, 2022", 3.52, 3.83, 3.22, 3.57, 3.57, 2151000],
    ["Jan 1, 2022", 4.77, 5.07, 3.04, 3.57, 3.57, 3416400],
    ["Dec 1, 2021", 6.23, 6.53, 4.58, 4.74, 4.74, 6459500],
    ["Nov 1, 2021", 6.91, 7.24, 5.92, 6.10, 6.10, 2325700],
    ["Oct 1, 2021", 6.69, 7.00, 6.45, 6.92, 6.92, 1466200],
    ["Sep 1, 2021", 7.12, 7.29, 6.66, 6.70, 6.70, 1963900],
    ["Aug 1, 2021", 6.78, 7.38, 6.56, 7.12, 7.12, 2607900],
    ["Jul 1, 2021", 7.18, 7.37, 6.55, 6.81, 6.81, 3222500],
    ["Jun 1, 2021", 6.78, 7.74, 6.54, 7.20, 7.20, 10532600],
    ["May 1, 2021", 7.63, 7.63, 6.10, 6.74, 6.74, 5320000],
    ["Apr 1, 2021", 8.27, 8.78, 7.45, 7.56, 7.56, 2422300],
    ["Mar 1, 2021", 8.19, 9.55, 7.80, 8.22, 8.22, 5140800],
    ["Feb 1, 2021", 7.25, 8.68, 6.68, 7.87, 7.87, 4753800],
    ["Jan 1, 2021", 7.86, 9.33, 6.26, 7.10, 7.10, 5815600],
    ["Dec 1, 2020", 6.93, 7.97, 6.51, 7.43, 7.43, 2545800],
    ["Nov 1, 2020", 6.15, 7.83, 5.78, 6.88, 6.88, 1759700],
    ["Oct 1, 2020", 5.73, 6.86, 5.15, 6.05, 6.05, 1567500],
    ["Sep 1, 2020", 5.60, 6.44, 5.13, 5.63, 5.63, 2626400],
    ["Aug 1, 2020", 5.21, 6.92, 5.15, 5.62, 5.62, 5465600],
    ["Jul 1, 2020", 4.52, 4.73, 3.95, 4.48, 4.48, 1930200],
    ["Jun 1, 2020", 4.72, 5.62, 4.34, 4.62, 4.62, 6823800],
    ["May 1, 2020", 4.76, 5.23, 3.56, 4.70, 4.70, 3435500],
    ["Apr 1, 2020", 3.45, 5.06, 3.35, 4.80, 4.80, 2114400],
    ["Mar 1, 2020", 7.07, 7.31, 2.55, 3.46, 3.46, 4350700],
    ["Feb 1, 2020", 7.49, 8.37, 6.95, 7.05, 7.05, 2058400],
    ["Jan 1, 2020", 6.63, 8.50, 6.34, 7.54, 7.54, 3152000],
    ["Dec 1, 2019", 6.00, 6.51, 5.86, 6.51, 6.51, 2247100]
]

# Create DataFrame
df = pd.DataFrame(data, columns=['Date', 'Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume'])

# Convert Date string to datetime
df['Date'] = pd.to_datetime(df['Date'])

# Create Excel writer object with xlsm format
writer = pd.ExcelWriter('stock_data_2.xlsm', engine='openpyxl')

# Write DataFrame to Excel
df.to_excel(writer, sheet_name='Stock Data', index=False)

# Get the workbook and worksheet
workbook = writer.book
worksheet = writer.sheets['Stock Data']

# Add formatting
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule

# Format headers
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Format number columns
for col in ['B', 'C', 'D', 'E', 'F']:  # Price columns
    for cell in worksheet[col]:
        if cell.row > 1:  # Skip header
            cell.number_format = '#,##0.00'

# Format volume column
for cell in worksheet['G']:
    if cell.row > 1:  # Skip header
        cell.number_format = '#,##0'

# Add color scaling to Close price column
worksheet.conditional_formatting.add('E2:E'+str(worksheet.max_row),
    ColorScaleRule(start_type='min', start_color='FF8080',
                  mid_type='percentile', mid_value=50, mid_color='FFEB84',
                  end_type='max', end_color='63BE7B'))

# Add borders
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

for row in worksheet.iter_rows():
    for cell in row:
        cell.border = thin_border

# Adjust column widths
for column in worksheet.columns:
    max_length = 0
    column = list(column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

# Save the workbook
workbook.save('stock_data_2.xlsx')